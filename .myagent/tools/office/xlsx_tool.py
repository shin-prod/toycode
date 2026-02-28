"""Excel (XLSX) 操作ツール。openpyxl を使用する。"""

import json
import os
from typing import Any, Optional

from utils.file_guard import check_write_allowed, resolve_path
from utils.logger import get_logger

logger = get_logger(__name__)


def xlsx_read(path: str, sheet_name: Optional[str] = None) -> str:
    """XLSX ファイルのセル内容を取得する。

    Args:
        path: XLSX ファイルのパス
        sheet_name: 読み込むシート名（省略時はアクティブシート）

    Returns:
        シートデータの JSON 文字列
    """
    try:
        import openpyxl
    except ImportError:
        return "エラー: openpyxl がインストールされていません。pip install openpyxl"

    resolved = resolve_path(path)
    if not os.path.isfile(resolved):
        return f"エラー: ファイルが存在しません: {resolved}"

    logger.debug("xlsx_read: %s sheet=%s", resolved, sheet_name)
    try:
        wb = openpyxl.load_workbook(resolved, read_only=True, data_only=True)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return (
                    f"エラー: シートが見つかりません: {sheet_name}\n"
                    f"利用可能なシート: {wb.sheetnames}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        rows_data = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append(list(row))

        # 末尾の空行を削除
        while rows_data and all(v is None for v in rows_data[-1]):
            rows_data.pop()

        result = {
            "sheet_name": sheet_name,
            "available_sheets": wb.sheetnames,
            "dimensions": ws.dimensions,
            "row_count": len(rows_data),
            "rows": rows_data,
        }
        return json.dumps(result, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        logger.error("xlsx_read エラー: %s", e)
        return f"エラー: XLSX 読み込み失敗: {e}"


def xlsx_write_cell(
    path: str, sheet_name: str, cell: str, value: Any
) -> str:
    """XLSX の指定セルに値を書き込んで保存する。

    Args:
        path: XLSX ファイルのパス
        sheet_name: 書き込み先シート名
        cell: セル番地 (例: A1, B3)
        value: 書き込む値

    Returns:
        成功メッセージまたはエラーメッセージ
    """
    try:
        import openpyxl
    except ImportError:
        return "エラー: openpyxl がインストールされていません。pip install openpyxl"

    resolved = resolve_path(path)
    if err := check_write_allowed(resolved):
        return err
    if not os.path.isfile(resolved):
        return f"エラー: ファイルが存在しません: {resolved}"

    logger.debug("xlsx_write_cell: %s sheet=%s cell=%s val=%s", resolved, sheet_name, cell, value)
    try:
        wb = openpyxl.load_workbook(resolved)
        if sheet_name not in wb.sheetnames:
            return (
                f"エラー: シートが見つかりません: {sheet_name}\n"
                f"利用可能なシート: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        ws[cell] = value
        wb.save(resolved)
        return f"セル {sheet_name}!{cell} に '{value}' を書き込みました: {resolved}"
    except Exception as e:
        logger.error("xlsx_write_cell エラー: %s", e)
        return f"エラー: XLSX 書き込み失敗: {e}"
